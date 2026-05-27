"""
src/loaders/excel_loader.py

Lee archivos .xlsx (diccionarios de datos) y retorna Documents con metadata
que incluye nombre del archivo, hoja y número de fila — esencial para la
citación de fuentes requerida en la rúbrica.
"""
from pathlib import Path
from typing import List

import openpyxl
from langchain_core.documents import Document

from src.config import config


def load_excel(file_path: str) -> List[Document]:
    """
    Procesa un archivo Excel completo.
    Cada hoja genera un Document por fila significativa (no vacía).
    Metadata: source, sheet_name, row_number, headers.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel no encontrado: {file_path}")

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    documents = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            continue

        # Primera fila = encabezados
        headers = [str(h).strip() if h is not None else f"col_{i}"
                   for i, h in enumerate(rows[0])]

        for row_idx, row in enumerate(rows[1:], start=2):  # start=2 por encabezado
            # Saltar filas completamente vacías
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            # Construir texto legible: "campo: valor | campo: valor ..."
            row_text = " | ".join(
                f"{headers[i]}: {cell}"
                for i, cell in enumerate(row)
                if i < len(headers) and cell is not None and str(cell).strip() != ""
            )

            if not row_text.strip():
                continue

            doc = Document(
                page_content=f"[{sheet_name}] Fila {row_idx}: {row_text}",
                metadata={
                    "source": file_path.name,
                    "full_path": str(file_path),
                    "sheet_name": sheet_name,
                    "row_number": row_idx,
                    "headers": ", ".join(headers),
                    "type": "data_dictionary",
                },
            )
            documents.append(doc)

    wb.close()
    print(f"[Excel] {len(documents)} filas cargadas desde {file_path.name}")
    return documents


def load_excel_folder(folder_path: str = None) -> List[Document]:
    """Carga todos los .xlsx de una carpeta."""
    folder = Path(folder_path or config.EXCEL_DIR)
    folder.mkdir(parents=True, exist_ok=True)

    all_docs = []
    for xlsx_file in folder.glob("*.xlsx"):
        all_docs.extend(load_excel(str(xlsx_file)))

    if not all_docs:
        print(f"[Excel] No se encontraron archivos .xlsx en {folder}")
    return all_docs
